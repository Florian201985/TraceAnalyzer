import pandas as pd
from openpyxl import load_workbook


def export_to_excel(filename_excel, parameters, trace_fft):

    x_key = 'X'
    y_key = 'Y'
    if trace_fft == 'fft':
        x_key = 'Frequencies'
        y_key = 'Amplitudes'

    # Find the maximum length of x and y across all parameters
    max_len_x = max(len(param[trace_fft][x_key]) for param in parameters)
    max_len_y = max(len(param[trace_fft][y_key]) for param in parameters)

    # Create a DataFrame
    df = pd.DataFrame()

    # Iterate through the parameters list
    for param in parameters:
        trace_param = param[trace_fft]
        x_label = trace_param['x_label']
        y_label = trace_param['y_label']
        if trace_fft == 'fft':
            x = list(trace_param[x_key])
            y = list(trace_param[y_key])
        else:
            x = list(trace_param[x_key].values)
            y = list(trace_param[y_key].values)



        for idx in range(max_len_x-len(x)):
            x.append(float('nan'))
            y.append(float('nan'))

        # Create a DataFrame for the current parameter
        df_param = pd.DataFrame({x_label: x, y_label: y})

        # Concatenate this parameter's DataFrame with the main DataFrame
        df = pd.concat([df, df_param], axis=1)

    # Write to Excel
    df.to_excel(filename_excel, index=False)

    # Open the created Excel file
    wb = load_workbook(filename_excel)

    # Select the active sheet
    ws = wb.active

    # Insert two rows at the beginning
    ws.insert_rows(1, amount=2)

    col = 1
    for param in parameters:
        # Set the values for filename and label
        ws.cell(row=1, column=col, value=param[trace_fft]['label'])
        ws.cell(row=2, column=col, value='')
        col += 2

    # Save the modified Excel file
    wb.save(filename_excel)